from VEDA_application.templatetags.veda_extras import get_avg_user, get_count_n, get_count_n_user
from django.dispatch.dispatcher import receiver

from django.shortcuts import render, HttpResponseRedirect, get_object_or_404

from django.contrib.auth.models import User
from django.contrib.auth import authenticate, logout, login
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, EmailMessage

from django.urls import reverse

from . import models
from VEDA_application.models import Admin_post, Client, Discipline, Group, Group_post, Lesson, Note, Notice, Notification, Parent, Task

import json
import openpyxl

from datetime import datetime


def index(request):
    """ Основная страница сайта, где расположены
        группы пользователя, их новости и блог разработчика """

   
    try:
        notes = Note.objects.filter(receiver__id=request.user.client.id)
        avg = get_avg_user(notes, request.user.client)
        count_n = get_count_n_user(notes, request.user.client)
        count_less_than_4 = 0
        for note in notes:
            try:
                if int(note.value) < 4:
                    count_less_than_4 += 1
            except:
                pass
    except:
        avg = 0
        count_n = 0
        count_less_than_4 = 4
    
    if request.method == 'GET':
        return render(request, 'index.html', context={
            'admin_posts': models.Admin_post.objects.all(), # посты от администратора
            'l_groups': models.Group.objects.filter(listeners__user__username=request.user.username),                     # группы, где есть текущий пользователь
            'm_groups': models.Group.objects.filter(moderators__user__username=request.user.username),                     # группы, где есть текущий пользователь
            'group_posts_listener': models.Group_post.objects.filter(group__listeners__user__username=request.user.username),    # посты в группах, в которых находится пользователь
            'group_posts_moderator': models.Group_post.objects.filter(group__moderators__user__username=request.user.username),    # посты в группах, в которых находится пользователь
            'notifications': models.Notification.objects.filter(receiver__user__username=request.user.username),        # уведомления для пользователя
            'tasks': models.Task.objects.filter(receiver__user__username=request.user.username),                       # задачи из группа для пользователя
            'avg': avg,
            'count_n': count_n,
            'count_less_than_4': count_less_than_4
        })


def logout_view(request):
    """ View для выхода из аккаунта авторизированного пользователя, в любом случае 
        (даже если не авторизирован пользователь) - возвращает на основную страницу """


    logout(request)
    return HttpResponseRedirect(reverse('index'))


def login_view(request):
    """ Страница входа в учетную запись пользователя, должна обрабатывать
        сохранение, вход и переходы на главную страницу, восстановление пароля
        и регистрацию пользователя """

    if request.method == 'GET':
        return render(request, 'login.html', context={

        })
    if request.method == "POST":
        user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'))
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse('index'))
        else:
            return render(request, 'login.html', context={
                'error': True
        })


def register(request):
    """ Страница регистрации, обрабатывает выдачу формы для ввода данных
        нового пользователя, а также принимает POST запрос на создание нового пользователя """

    if request.method == "GET":
        return render(request, 'register.html', context={

        })
    if request.method == "POST":
        try:
            user = User.objects.create_user(username=request.POST.get('username'), 
                email=request.POST.get('email'), 
                first_name=request.POST.get('firstname'), 
                last_name=request.POST.get('secondname'), 
                password=request.POST.get('password'))
                
            # user.client.adress = request.POST.get('address')
            # user.client.hobbies = request.POST.get('hobbies')
            user.client.day_of_birthday = request.POST.get('dob')
            user.client.gender = request.POST.get('gender')
            user.client.status = ""

            user.save()

            if user is not None:
                login(request, user)

                send_mail('Регистрация', 
                    f'Поздравляем, вы зарегистрировались на нашей платформе, ваш логин: {user.username}! Приятного использования!',
                    'labark.group@gmail.com',
                    [user.email])
                return HttpResponseRedirect(reverse('index'))
        except Exception as e:
            print(e)
            return render(request, 'register.html', context={
                'error': e
            })
    

def profile(request, pk):
    """ Отдает страницу профиля, и, в случае, если это профиль пользователя, то может принять
        UPDATE запрос на изменение данных профиля """

    if request.method == "GET":
        client = get_object_or_404(models.Client, pk=pk)
        groups = models.Group.objects.filter(listeners__user__username=client.user.username)

        parents_access = False
        for group in groups:
            if request.user.client in group.moderators.all():
                parents_access = True

        return render(request, 'profile.html', context={
            'client': client,
            'groups': groups,
            'parents_access': parents_access
        })


def post(request):
    """ Страница конкретного поста блога (как группы, так и разработчика), может быть POST,
        тогда создает новый пост в области создания """

    if request.method == "GET":
        return render(request, 'post.html', context={

        })
    if request.method == "POST":
        if request.user.is_staff:
            newPost = Admin_post(author=request.user.client, article=request.POST.get('article'), text=request.POST.get('content'))
            newPost.save()
        return HttpResponseRedirect(reverse('index'))  


def group(request, pk):
    """ Страница, отображающая всю основную информацию о комнате группы, принимает POST,
        чтобы создать страницу группы, UPDATE для изменения данных """

    if request.method == "GET":
        group = get_object_or_404(models.Group, pk=pk)

        listeners = group.listeners.all()
        moderators = group.moderators.all()
        notes = Note.objects.filter(group__id=group.id)

        sorted_users = list(set(listeners | moderators))
        sorted_users = sorted(sorted_users, key=lambda user: -get_avg_user(notes, user))

        return render(request, 'group.html', context={
            'group': group,
            'listeners': listeners,
            'moderators': moderators,
            'sorted_users': sorted_users,
            'notices': models.Notice.objects.filter(group__id=group.id),
            'tasks': models.Task.objects.filter(group__id=group.id),
            'posts': models.Group_post.objects.filter(group__id=group.id),
            'sheet': models.Lesson.objects.filter(group__id=group.id),
            'disciplines': models.Discipline.objects.filter(group__id=group.id),
            'days': range(1, 32),
            'notes': notes,
            'months': [
                (1, 'Январь'),
                (2, 'Февраль'),
                (3, 'Март'),
                (4, 'Апрель'),
                (5, 'Май'),
                (6, 'Июнь'),
                (7, 'Июль'),
                (8, 'Август'),
                (9, 'Сентябрь'),
                (10, 'Октябрь'),
                (11, 'Ноябрь'),
                (12, 'Декабрь'),
            ],
        })


@login_required
def create_group(request):
    """ Страница создания комнаты группы """

    if request.method == "GET":
        return render(request, 'create_group.html', context={

        })
    elif request.method == "POST":
        data = json.loads(request.body)

        group = Group.objects.create(
            name = data['name'],
            tag = data['tag'],
            hNotices = data['hNotices'],
            hTasks = data['hTasks'],
            hSheet = data['hSheet'],
        )

        group.moderators.add(request.user.client)
        group.save()

        notification = Notification.objects.create(
            receiver=request.user.client,
            message=f'Вы создали группу {group.name}',
            priority=2
        )

        disciplines = []

        for lmo in data['mo']:
            if lmo not in disciplines:
                disciplines.append(lmo)

            lesson = Lesson(
                group=group,
                discipline=lmo,
                day=1,
            )
            
            lesson.save()
        
        for lt in data['tu']:
            if lt not in disciplines:
                disciplines.append(lt)

            lesson = Lesson(
                group=group,
                discipline=lt,
                day=2
            )
            
            lesson.save()

        for lwe in data['we']:
            if lwe not in disciplines:
                disciplines.append(lwe)

            lesson = Lesson(
                group=group,
                discipline=lwe,
                day=3
            )
            
            lesson.save()
   
        for lth in data['th']:
            if lth not in disciplines:
                disciplines.append(lth)

            lesson = Lesson(
                group=group,
                discipline=lth,
                day=4
            )
            
            lesson.save()
        
        for lfr in data['fr']:
            if lfr not in disciplines:
                disciplines.append(lfr)

            lesson = Lesson(
                group=group,
                discipline=lfr,
                day=5
            )
            
            lesson.save()

        for lsa in data['sa']:
            if lsa not in disciplines:
                disciplines.append(lsa)

            lesson = Lesson(
                group=group,
                discipline=lsa,
                day=6
            )
            
            lesson.save()
        
        for lsu in data['su']:
            if lsu not in disciplines:
                disciplines.append(lsu)

            lesson = Lesson(
                group=group,
                discipline=lsu,
                day=7
            )
            
            lesson.save()
        
        for discipline in disciplines:
            new_discipline = Discipline.objects.create(
                group = group,
                name = discipline
            )

        return HttpResponseRedirect(reverse('group', args=[group.id]))


@login_required
def set_lesson(request, pk):
    """ Вью для обновления значения урока в расписании """

    if request.method == "POST":
        lesson = get_object_or_404(Lesson, pk=pk)

        new_value = request.POST.get('discipline')

        if len(new_value) > 0:
            lesson.discipline = new_value
            lesson.save()
        else:
            lesson.delete()

        return HttpResponseRedirect(reverse('group', args=[lesson.group.id]))


@login_required
def add_lesson(request, pk):
    """ Вью для добавления урока в расписание, рк - ключ группы, остальное -- в запрсое """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        lesson = Lesson.objects.create(
            group = group,
            day = request.POST.get('day'),
            discipline = request.POST.get('discipline')
        )

        for moderator in group.moderators.all():
            notification = Notification.objects.create(
                receiver = moderator,
                message = f"В группе {group.name} обновлено расписание!",
                priority = 4
            )

        for listener in group.listeners.all():
            notification = Notification.objects.create(
                receiver = listener,
                message = f"В группе {group.name} обновлено расписание!",
                priority = 4
            )

        return HttpResponseRedirect(reverse('group', args=[group.id]))


def add_discipline(request, pk):
    """ Вью для создания предмета для ведения ведомости """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        discipline = Discipline.objects.create(
            group = group,
            name = request.POST.get('name')
        )

        return HttpResponseRedirect(reverse('group', args=[pk]))


def remove_discipline(request, pk):
    """ Вью для удаления предмета для ведения ведомости """

    if request.method == "POST":
        discipline = get_object_or_404(Discipline, pk=pk)

        discipline.delete()

        return HttpResponseRedirect(reverse('group', args=[discipline.group.id]))


def group_enter(request, pk):
    """ Вью для входа в группу конкретного пользователя в группу под РК """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        group.listeners.add(request.user.client)
        group.save()

        moderators = group.moderators.all()

        for moderator in moderators:
            notification = Notification.objects.create(
                receiver = moderator,
                message = f'@{request.user.username} зашел в группу #{group.name}',
                priority = 2
            )

        return HttpResponseRedirect(reverse('index'))


def group_out(request, pk):
    """ Вью для выхода пользователя из группы """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)
        
        if request.user.client in group.listeners.all():
            group.listeners.remove(request.user.client)
        elif request.user.client in group.moderators.all():
            group.moderators.remove(request.user.client)
        
        group.save()

        moderators = group.moderators.all()

        for moderator in moderators:
            notification = Notification.objects.create(
                receiver = moderator,
                message = f'@{request.user.username} вышел из группы #{group.name}',
                priority = 2
            )

        return HttpResponseRedirect(reverse('index'))


def group_post(request, pk):
    """ Вью для создания постов в группе """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        group_post = Group_post.objects.create(
            group = group,
            article = request.POST.get('article'),
            text = request.POST.get('text'),
            author = request.user.client,
        )

        try:
            file = request.FILES['file']

            if file:
                group_post.post_file_field = file
                group_post.save()
        except:
            pass

        for listener in group.listeners.all():
            notification = Notification.objects.create(
                receiver = listener,
                message = f'В группе #{group.name} новый пост!',
                priority = 4
            )
        
        for moderator in group.moderators.all():
            notification = Notification.objects.create(
                receiver = moderator,
                message = f'В группе #{group.name} новый пост!',
                priority = 4
            )
        
        return HttpResponseRedirect(reverse('group', args=[pk]))


def get_results(request, pk):
    """ Получение на почту ведомости группы """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        wb = openpyxl.Workbook()
        wb.create_sheet(title = 'Ведомость', index = 0)

        sheet = wb['Ведомость']
        max_count_notes = 0

        notes = Note.objects.filter(group__id=group.id)
        notes_dict = {}
        for note in notes:
            if note.receiver.id not in notes_dict.keys():
                notes_dict[note.receiver.id] = []
                notes_dict[note.receiver.id].append(note.value)
            else:
                notes_dict[note.receiver.id].append(note.value)
        
        cell = sheet.cell(1,1)
        cell.value = f'Ведомость группы {group.name} за 12 месяцев'

        iterator = 3
        for key in notes_dict.keys():
            receiver = get_object_or_404(Client, pk=key)

            cell = sheet.cell(iterator, 1)
            cell.value = receiver.user.first_name + " " + receiver.user.last_name

            if max_count_notes < len(notes_dict[key]):
                max_count_notes = len(notes_dict[key])

            col_iterator = 2

            for note in notes_dict[key]:
                cell = sheet.cell(iterator, col_iterator)
                cell.value = note
                
                col_iterator += 1
            
            iterator += 1

        iterator = 3
        col = max_count_notes + 2

        cell = sheet.cell(iterator - 1, col)
        cell.value = 'Средний балл'

        cell = sheet.cell(iterator - 1, col + 1)
        cell.value = 'Количество пропусков'

        for key in notes_dict.keys():
            cell = sheet.cell(iterator, col)
            cell.value = get_avg_user(notes, get_object_or_404(Client, pk=key ))

            cell = sheet.cell(iterator, col + 1)
            cell.value = get_count_n_user(notes, get_object_or_404(Client, pk=key ))

            iterator += 1

        wb.save(f'Ведомость группы.xlsx')

        mail = EmailMessage('Ваша ведомость группы!', 
            "Мы собрали для вас ведомость группы в прикрепленном файле!", 
            "labark.group@gmail.com",
            [request.user.email])

        with open('Ведомость группы.xlsx', 'rb') as report:
            mail.attach('Ведомость группы.xlsx', report.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            res = mail.send()

        return HttpResponseRedirect(reverse('group', args=[group.id]))



@login_required
def group_search(request):
    """ Вью для поиска группы по идентификтору из буков """

    if request.method == "GET":
        group = get_object_or_404(Group, tag=request.GET.get('tag'))

        return HttpResponseRedirect(reverse('group', args=[group.id]))


@login_required
def update_status_group(request, pk):
    """ Вью для обновления статуса между модератором и слушателем в группе """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        client = get_object_or_404(Client, pk=request.POST.get('user'))

        if client in group.moderators.all():
            group.moderators.remove(client)
            group.listeners.add(client)
        elif client in group.listeners.all():
            group.listeners.remove(client)
            group.moderators.add(client)

        group.save()

        return HttpResponseRedirect(reverse('group', args=[pk]))


@login_required
def kick_user(request, pk):
    """ Вью для обновления статуса между модератором и слушателем в группе """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk)

        client = get_object_or_404(Client, pk=request.POST.get('user'))

        if client in group.listeners.all():
            group.listeners.remove(client)

        group.save()

        return HttpResponseRedirect(reverse('group', args=[pk]))


def faq(request):
    """ Страница ответов на основные вопросы к проекту """

    if request.method == "GET":
        return render(request, 'faq.html', context={

        })


def settings(request):
    """ Страница просмотра и изменения настроек пользователя """

    if request.method == "GET":
        return render(request, 'settings.html', context={

        })


def update_user(request, pk):
    """ Вью для обновления настроек и личной информации пользователя """

    if request.method == "POST":
        user = get_object_or_404(User, pk=pk)

        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')

        user.client.status = request.POST.get('status')
        user.client.hobbies = request.POST.get('hobbies')
        user.client.adress = request.POST.get('adress')

        user.client.vk = request.POST.get('vk')
        user.client.instagram = request.POST.get('instagram')

        user.save()

        return HttpResponseRedirect(reverse('settings'))


@login_required
def update_password(request, pk):
    """ Вью для изменения пароля пользователя """

    if request.method == "POST":
        user = get_object_or_404(User, pk=pk)

        if (user.check_password(request.POST.get('password'))):
            user.set_password(request.POST.get('new_password'))
            user.save()

            user = authenticate(username=user.username, password=request.POST.get('new_password'))
            if user is not None:
                login(request, user)
                return HttpResponseRedirect(reverse('settings'))
            else:
                return render(request, 'login.html', context={
                    'error': True
            })
        else:
            return render(request, 'settings.html', context={
                'password_error': True
            })  


def update_father(request):
    """ Вью для изменения информации о отце """

    if request.method == "POST":
        try:
            parent = request.user.client.father

            parent.surname = request.POST.get('surname')
            parent.name = request.POST.get('name')
            parent.middlename = request.POST.get('middlename')
            parent.job = request.POST.get('job')
            parent.dob = request.POST.get('dob')
            parent.phone = request.POST.get('phone')

            parent.save()
        except:
            parent = Parent()

            parent.surname = request.POST.get('surname')
            parent.name = request.POST.get('name')
            parent.middlename = request.POST.get('middlename')
            parent.job = request.POST.get('job')
            parent.dob = request.POST.get('dob')
            parent.phone = request.POST.get('phone')

            parent.save()

            request.user.client.father = parent
            request.user.save()

        return HttpResponseRedirect(reverse('settings'))


def update_mother(request):
    """ Вью для изменения информации о отце """

    if request.method == "POST":
        try:
            parent = request.user.client.mother

            parent.surname = request.POST.get('surname')
            parent.name = request.POST.get('name')
            parent.middlename = request.POST.get('middlename')
            parent.job = request.POST.get('job')
            parent.dob = request.POST.get('dob')
            parent.phone = request.POST.get('phone')

            parent.save()
        except:
            parent = Parent()

            parent.surname = request.POST.get('surname')
            parent.name = request.POST.get('name')
            parent.middlename = request.POST.get('middlename')
            parent.job = request.POST.get('job')
            parent.dob = request.POST.get('dob')
            parent.phone = request.POST.get('phone')

            parent.save()

            request.user.client.mother = parent
            request.user.save()

        return HttpResponseRedirect(reverse('settings'))


def update_userpic(request):
    """ Вью для изменения аватарки пользователя """

    if request.method == "POST":
        # try:
            user = request.user

            user.client.userpic = request.FILES['file']
            user.save()

            return HttpResponseRedirect(reverse('settings'))
        # except Exception as e:
        #     print(e)
            
            return HttpResponseRedirect(reverse('index'))


def recover(request):
    """ Страница восстановления пароля пользователя """

    if request.method == "GET":
        return render(request, 'recover.html', context={

        })


def add_note(request, pk_group, pk_user):
    """ Вью для добавления оценки в группу """

    if request.method == "POST":
        group = get_object_or_404(Group, pk=pk_group)
        receiver = get_object_or_404(Client, pk=pk_user)

        date_of_create = request.POST.get('date')
        discipline = get_object_or_404(Discipline, pk=request.POST.get('discipline')) 
        value = request.POST.get('value')

        description = request.POST.get('description')

        note = Note.objects.create(
            group = group,
            receiver = receiver,
            date_of_receive = date_of_create,
            value = value,
            discipline = discipline,
            description = description
        )

        notification = Notification.objects.create(
            receiver = receiver,
            message = f'Вы получили оценку {note.value} в группе {group.name}',
            priority = 1
        )

        print(receiver)
        print(group)
        print(pk_user)
        print(pk_group)

        return HttpResponseRedirect(reverse('group', args=[group.id]))


def remove_note(request, pk):
    """ Вью для удаления конкретной оценки """

    if request.method == "POST":
        note = get_object_or_404(Note, pk=pk)

        note.delete()

        return HttpResponseRedirect(reverse('group', args=[note.group.id]))


@login_required
def add_task(request, pk):
    """ Вью для добавления задачи пользователю в конкретной группе """

    if request.method=="POST":
        group = get_object_or_404(Group, pk=pk)
        user = get_object_or_404(Client, pk=request.POST.get('user'))

        task = Task.objects.create(
            group = group,
            receiver = user,
            content = request.POST.get('content'),
            status = 0
        )

        notification = Notification.objects.create(
            receiver = user,
            message = f"Вам пришла новая задача в группе #{group.name}",
            priority = 3
        )

        return HttpResponseRedirect(reverse('group', args=[pk]))


@login_required
def update_task(request, pk):
    """ Вью для изменения статуса выполнения задачи в конкретной группе """

    if request.method == "POST":
        task = get_object_or_404(Task, pk=pk)

        if task.status == 0:
            task.status = 1
        else:
            task.status = 0
        
        task.date_of_completion = datetime.now()

        task.save()

        return HttpResponseRedirect(reverse('group', args=[task.group.id]))


@login_required
def remove_task(request, pk):
    """  Вью для удаления задачи из конкретной группы """

    pass


@login_required
def remove_notification(request, pk):
    """ Вьюшка для удаления уведомления """

    if request.method == "POST":
        notification = get_object_or_404(Notification, pk=pk)
        notification.delete()

        return HttpResponseRedirect(reverse('index'))  


@login_required
def add_notice(request, pk):
    """ Вьюшка для добавления уведомления """

    if request.method == "POST":
        group = get_object_or_404(models.Group, pk=pk)
        notice = models.Notice(group=group, author=request.user.client, message=request.POST.get('message'))
        notice.save()

        for listener in notice.group.listeners.all():
            notification = Notification.objects.create(
                receiver = listener,
                message = f'Новое объявление в группе {group.name}',
                priority = 4
            )

        for moderator in notice.group.moderators.all():
            notification = Notification.objects.create(
                receiver = moderator,
                message = f'Новое объявление в группе {group.name}',
                priority = 4
            )

        return HttpResponseRedirect(reverse('group', args=[group.id]))  


@login_required
def remove_notice(request, pk):
    """ Вьюшка для удаления уведомления """

    if request.method == "POST":
        notice = get_object_or_404(Notice, pk=pk)
        group = get_object_or_404(Group, pk=notice.group.id)
        
        notice.delete()

        return HttpResponseRedirect(reverse('group', args=[group.id]))  
